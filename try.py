import os
from openpyxl import load_workbook,Workbook
from xlrd import open_workbook
from xlutils.copy import copy    
wb = Workbook()
ws = wb.active
if not os.path.exists("/home/poojahegde/newtry/STLab/student.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name","USN","Marks1","Marks2","Marks3"])
else:
    wb1 = open_workbook('/home/poojahegde/newtry/STLab/student.xlsx')
    # ws = wb1.get_sheet(0)
    ws = wb1.active
    
    
for i in range(2):
    print("Enter Name, USN, Marks1, Marks2, Marks3 for student")
    data = input().split(" ")
    ws.append([data[0],data[1],data[2],data[3],data[4]])
wb.save("student.xlsx")
os.system("student.xlsx")


