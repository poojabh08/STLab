import re
import openpyxl
from openpyxl import Workbook
key1 = input("Enter sixteen character License Key: ")
wb = openpyxl.load_workbook("/home/poojahegde/newtry/STLab/LicenseKey.xlsx")
ws = wb.active
rows = ws.max_row
columns = ws.max_column
f=1
for i in range(1,rows+1):
    cell_obj = ws.cell(row=i,column=2)
    key = cell_obj.value
    res = re.search(key1,key)
    if res:
        f=0
        print(ws.cell(row=i,column=1).value,"           ",res.group())
if f==1:
        print("License key Not Found!!!!!:(")
  
