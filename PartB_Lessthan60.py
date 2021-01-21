import os
import openpyxl
from openpyxl import Workbook 
def updateAndWriteNew():
    wb = openpyxl.load_workbook("student.xlsx")
    ws = wb.active
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Name","USN","Marks 1","Marks 2","Marks 3"])
    rows = ws.max_row
    for i in range(2,rows+1):
        for j in range(3,6):
            cell_obj = ws.cell(row=i,column=j)
            if int(cell_obj.value) < 60:
                lst=[]         
                for k in range(1,6):
                    lst.append(ws.cell(row=i,column=k).value)
                print(lst)
                ws2.append([lst[0],lst[1],lst[2],lst[3],lst[4]])
                wb2.save("spreadsheet2.xlsx")
                break

if __name__=="__main__":    
    updateAndWriteNew()
    os.system("spreadsheet2.xlsx")